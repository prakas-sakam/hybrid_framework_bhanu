import openpyxl

def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_colunm_count(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column

def get_cell_data(path, sheet_name, row_number, col_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=col_number).value

def get_cell_data(path, sheetname, row_number, col_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=row_number, column=col_number).value=data
    workbook.save(path)

def get_data_from_excel(path, sheet_name):

    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(2, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
    return final_list
